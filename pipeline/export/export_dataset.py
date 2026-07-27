"""Export the qualified records to the deliverable: a multi-sheet XLSX + a CSV.

Sheets:
  1. Dataset                 — clean, customer-facing cells only.
  2. Provenance & Verification — the basis behind every high-value cell.
  3. Audit (rejected/removed) — firms that failed Rule 2 and cell values pulled in validation.
  4. Sources & Method        — discovery source-class catalogue + concentration check.

Selection: only status='qualified' records; capped so no single discovery source class
exceeds SOURCE_CAP of the final set (defends the single-source rule); filled to TARGET
preferring higher completeness tiers and single-family offices.
"""
from __future__ import annotations
import sys, os, csv
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))
from pipeline.common import store
from pipeline.common.schema import CUSTOMER_FACING_FIELDS, PROVENANCE_FIELDS, DISCOVERY_SOURCE_CLASSES

TARGET = 50
SOURCE_CAP_FRAC = 0.34
OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "data", "final")

_TIER_RANK = {"A+": 0, "A": 1, "B+": 2, "B": 3, "C+": 4, "C": 5, "": 6}


def _rank(r):
    # lower is better: higher tier, SFO first (rarer/more valuable), has email
    return (_TIER_RANK.get(r.get("completeness_tier", ""), 6),
            0 if r.get("fo_type") == "SFO" else 1,
            0 if r.get("principal_email") else 1)


def select():
    qualified = [r for r in store.all_candidates() if r.get("status") == "qualified"]
    qualified.sort(key=_rank)
    cap = max(1, int(TARGET * SOURCE_CAP_FRAC))
    chosen, per_class = [], {}
    # first pass: respect the per-class cap
    for r in qualified:
        cls = r.get("discovery_source_class", "?")
        if per_class.get(cls, 0) >= cap:
            continue
        chosen.append(r); per_class[cls] = per_class.get(cls, 0) + 1
        if len(chosen) >= TARGET:
            break
    return chosen, per_class, len(qualified)


def _autofit(ws):
    from openpyxl.utils import get_column_letter
    for i, col in enumerate(ws.columns, 1):
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 12), 60)


def export():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    os.makedirs(OUT_DIR, exist_ok=True)
    chosen, per_class, n_qual = select()

    wb = openpyxl.Workbook()
    head_fill = PatternFill("solid", fgColor="1F3A5F")
    head_font = Font(color="FFFFFF", bold=True)

    def write_sheet(ws, title, fields, rows):
        ws.title = title
        ws.append([f.replace("_", " ").title() for f in fields])
        for c in ws[1]:
            c.fill = head_fill; c.font = head_font; c.alignment = Alignment(vertical="top", wrap_text=True)
        for r in rows:
            ws.append([r.get(f, "") for f in fields])
        ws.freeze_panes = "A2"
        _autofit(ws)

    write_sheet(wb.active, "Dataset", CUSTOMER_FACING_FIELDS, chosen)
    write_sheet(wb.create_sheet(), "Provenance & Verification", PROVENANCE_FIELDS, chosen)

    audit_rows = store.connect().execute(
        "SELECT ts,firm,field,value,reason FROM audit ORDER BY firm").fetchall()
    aws = wb.create_sheet("Audit (rejected & removed)")
    aws.append(["Date", "Firm", "Field/Decision", "Rejected value", "Reason"])
    for c in aws[1]:
        c.fill = head_fill; c.font = head_font
    for a in audit_rows:
        aws.append([a["ts"], a["firm"], a["field"], a["value"], a["reason"]])
    aws.freeze_panes = "A2"; _autofit(aws)

    sws = wb.create_sheet("Sources & Method")
    sws.append(["Discovery source class", "In final 50", "% of final"])
    for c in sws[1]:
        c.fill = head_fill; c.font = head_font
    total = max(len(chosen), 1)
    for cls in DISCOVERY_SOURCE_CLASSES:
        n = per_class.get(cls, 0)
        if n:
            sws.append([cls, n, f"{100*n/total:.0f}%"])
    sws.append([])
    sws.append(["Total qualified in store", n_qual, ""])
    sws.append(["Selected into final", len(chosen), ""])
    _autofit(sws)

    xlsx = os.path.join(OUT_DIR, "family_office_dataset.xlsx")
    wb.save(xlsx)

    csv_path = os.path.join(OUT_DIR, "family_office_dataset.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(CUSTOMER_FACING_FIELDS)
        for r in chosen:
            w.writerow([r.get(k, "") for k in CUSTOMER_FACING_FIELDS])

    print(f"Exported {len(chosen)} records -> {xlsx}")
    print(f"Source mix in final: {per_class}")
    print(f"Qualified available: {n_qual}")
    return chosen, per_class


if __name__ == "__main__":
    export()
